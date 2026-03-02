from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.schema import TestSuite
from core.config import config


def _validate_output_path(output_path: str) -> str:
    """
    Validate that the output path is within the configured output directory.
    Prevents path traversal attacks (e.g., writing to ../../.ssh/).
    """
    # Resolve output_dir relative to project root, not CWD
    project_root = Path(__file__).parent.parent
    path = Path(output_path)
    if not path.is_absolute():
        abs_path = (project_root / path).resolve()
    else:
        abs_path = path.resolve()
    allowed_dir = (project_root / config.output.output_dir).resolve()
    
    try:
        abs_path.relative_to(allowed_dir)
    except ValueError:
        raise PermissionError(
            f"Output path must be within '{config.output.output_dir}/'. "
            f"Got: {output_path}"
        )
    
    return str(abs_path)


def _join_lines(items: List[str]) -> str:
    return "\n".join(items) if items else ""


def _export_excel(suite: TestSuite, output_path: str) -> None:
    workbook = Workbook()

    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet(title="Test Cases")
    else:
        sheet.title = "Test Cases"

    headers = [
        "Use Case",
        "Test Case",
        "Test Type",
        "Preconditions",
        "Test Data",
        "Steps",
        "Priority",
        "Dependencies",
        "Tags",
        "Expected Results",
        "Actual Results",
    ]

    sheet.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True)
        sheet.column_dimensions[get_column_letter(col_idx)].width = 30

    for test_case in suite.test_cases:
        sheet.append([
            test_case.use_case,
            test_case.test_case,
            getattr(test_case, 'test_type', 'functionality'),
            _join_lines(test_case.preconditions),
            "\n".join(f"{k}: {v}" for k, v in test_case.test_data.items()),
            _join_lines(test_case.steps),
            test_case.priority,
            ", ".join(test_case.dependencies),
            ", ".join(test_case.tags),
            _join_lines(test_case.expected_results),
            _join_lines(test_case.actual_results or []),
        ])

    # no need to delete default sheet when using workbook.active

    validated = _validate_output_path(output_path)
    Path(validated).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(validated)


def _export_json(suite: TestSuite, output_path: str) -> None:
    validated = _validate_output_path(output_path)
    Path(validated).parent.mkdir(parents=True, exist_ok=True)
    with open(validated, "w", encoding="utf-8") as f:
        f.write(suite.model_dump_json(indent=2))


def export(
    suite: TestSuite,
    output_format: str,
    output_path: str,
) -> None:
    fmt = output_format.lower().strip()
    if fmt == "excel":
        _export_excel(suite, output_path)
    elif fmt == "json":
        _export_json(suite, output_path)
    else:
        raise ValueError(f"Unsupported output format: {output_format}")
